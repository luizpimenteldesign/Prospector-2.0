# ========== IMPORTS / CONFIG ==========
import streamlit as st
import pandas as pd
from urllib.parse import quote
import requests
from datetime import datetime
import time
from ibge_localidades import buscar_estados, buscar_cidades_por_estado
from nichos_comerciais import obter_todos_nichos, obter_tags_osm_nicho, obter_categorias_nicho

st.set_page_config(
    page_title="Agente de Prospecção | LP Design",
    page_icon="🧭",
    layout="wide",
)

st.markdown("""
<style>
    .stApp { background-color: #0e1117; }
    [data-testid="stSidebar"] { background-color: #1a1d24; }
    .stButton > button {
        background: linear-gradient(90deg, #ff6b35, #ff8c42);
        color: white;
        border: none;
        border-radius: 8px;
        font-weight: 600;
    }
    .stButton > button:hover {
        background: linear-gradient(90deg, #ff8c42, #ff6b35);
        box-shadow: 0 4px 12px rgba(255, 107, 53, 0.4);
    }
    [data-testid="stMetricValue"] { color: #ff6b35; font-size: 28px; }
</style>
""", unsafe_allow_html=True)

st.logo(
    "https://luizpimentel.com/wp-content/uploads/2022/09/PNG-72dpi-Identidade-Visual-LP-Design-05.png",
    size="large",
)


# ========== FUNÇÕES ==========
def buscar_logo_site(url):
    if not url:
        return None
    try:
        from urllib.parse import urlparse
        parsed = urlparse(url if url.startswith('http') else 'https://' + url)
        domain = parsed.netloc or parsed.path
        return f"https://www.google.com/s2/favicons?domain={domain}&sz=128"
    except:
        return None


def analisar_site(url):
    if not url:
        return {"responde": False, "tem_https": False, "tem_mobile": False, "wordpress": False, "tempo": 0}
    
    try:
        if not url.startswith('http'):
            url = 'https://' + url
        
        start = time.time()
        response = requests.get(url, timeout=5, allow_redirects=True)
        tempo = time.time() - start
        html = response.text.lower()
        
        return {
            "responde": response.status_code == 200,
            "tem_https": url.startswith('https'),
            "tem_mobile": 'viewport' in html,
            "wordpress": 'wp-content' in html or 'wp-includes' in html,
            "tempo": round(tempo, 2)
        }
    except:
        return {"responde": False, "tem_https": False, "tem_mobile": False, "wordpress": False, "tempo": 0}


def calcular_prioridade_score(lead_data, analise_site):
    score = 0
    sugestoes = []
    
    if not lead_data.get("site"):
        score += 40
        sugestoes.append("🌐 Criação de Site")
    elif not analise_site["responde"]:
        score += 35
        sugestoes.append("🔧 Site Offline")
    else:
        if not analise_site["tem_https"]:
            score += 15
            sugestoes.append("🔒 HTTPS")
        if analise_site["tempo"] > 3:
            score += 15
            sugestoes.append("⚡ Performance")
        if not analise_site["tem_mobile"]:
            score += 20
            sugestoes.append("📱 Mobile")
    
    if not lead_data.get("facebook") and not lead_data.get("instagram"):
        score += 20
        sugestoes.append("📱 Redes Sociais")
    
    sugestoes.append("🎨 Identidade Visual")
    sugestoes.append("📊 Marketing Digital")
    
    if score >= 70:
        prioridade = "🔴 Alta"
    elif score >= 40:
        prioridade = "🟡 Média"
    else:
        prioridade = "🟢 Baixa"
    
    return {
        "prioridade": prioridade,
        "score": min(score, 100),
        "sugestoes": sugestoes[:5]
    }


def gerar_mensagem_whatsapp(empresa, cidade):
    return f"Olá! Encontrei {empresa} em {cidade} e vejo oportunidades de melhorar a presença digital. Sou da LP Design. Podemos conversar?"


def geocodificar_cidade(cidade, estado):
    try:
        url = "https://nominatim.openstreetmap.org/search"
        params = {"q": f"{cidade}, {estado}, Brasil", "format": "json", "limit": 1}
        headers = {"User-Agent": "LP-Design-Prospector/2.0"}
        time.sleep(1)
        response = requests.get(url, params=params, headers=headers, timeout=10)
        data = response.json()
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
        return -15.7939, -47.8828
    except:
        return -15.7939, -47.8828


def mapear_categoria_para_tags(categoria):
    """Mapeia categorias específicas para tags OSM precisas"""
    mapeamento = {
        # Alimentação
        "restaurantes": ["amenity=restaurant"],
        "cafeterias": ["amenity=cafe"],
        "lanchonetes e fast-food": ["amenity=fast_food"],
        "pizzarias": ["amenity=restaurant", "cuisine=pizza"],
        "padarias": ["shop=bakery"],
        "bares e pubs": ["amenity=bar", "amenity=pub"],
        "sorveterias": ["shop=ice_cream", "amenity=ice_cream"],
        "confeitarias": ["shop=confectionery", "shop=pastry"],
        
        # Saúde
        "clínicas médicas": ["amenity=clinic", "amenity=doctors"],
        "consultórios odontológicos": ["amenity=dentist"],
        "farmácias": ["amenity=pharmacy"],
        "academias": ["leisure=fitness_centre", "leisure=sports_centre"],
        "fisioterapia": ["amenity=clinic", "healthcare=physiotherapist"],
        
        # Beleza
        "salões de beleza": ["shop=beauty", "shop=hairdresser"],
        "barbearias": ["shop=barber"],
        
        # Serviços
        "escritórios de advocacia": ["office=lawyer"],
        "contabilidade": ["office=accountant"],
        "imobiliárias": ["office=estate_agent"],
        
        # Varejo
        "lojas de roupas": ["shop=clothes"],
        "calçados": ["shop=shoes"],
        "eletrônicos": ["shop=electronics"],
        "móveis e decoração": ["shop=furniture"],
        
        # Automotivo
        "oficinas mecânicas": ["shop=car_repair"],
        "lava-jatos": ["amenity=car_wash"],
    }
    
    categoria_lower = categoria.lower()
    for key in mapeamento:
        if key in categoria_lower:
            return mapeamento[key]
    
    return None


def buscar_leads_overpass(cidade, estado, max_leads, nicho, categoria):
    try:
        lat, lon = geocodificar_cidade(cidade, estado)
        
        # Tenta mapear categoria específica primeiro
        if categoria != "Todas":
            tags = mapear_categoria_para_tags(categoria)
            if not tags:
                # Se não encontrou mapeamento, usa tags do nicho
                tags = obter_tags_osm_nicho(nicho)
        else:
            tags = obter_tags_osm_nicho(nicho)
        
        if not tags:
            tags = ["shop"]
        
        raio_metros = 20000
        
        query = f"[out:json][timeout:30];("
        for tag in tags:
            key, value = tag.split("=")
            query += f'node["{key}"="{value}"](around:{raio_metros},{lat},{lon});'
            query += f'way["{key}"="{value}"](around:{raio_metros},{lat},{lon});'
        query += ");out center;"
        
        url = "https://overpass-api.de/api/interpreter"
        
        with st.spinner(f"🔍 Buscando em {cidade}/{estado}..."):
            time.sleep(2)
            response = requests.post(url, data={"data": query}, timeout=40)
            data = response.json()
        
        elements = data.get("elements", [])[:max_leads]
        
        leads = []
        for i, element in enumerate(elements, 1):
            tags = element.get("tags", {})
            
            nome = tags.get("name", f"Estabelecimento {i}")
            telefone = tags.get("phone", tags.get("contact:phone", ""))
            whatsapp = "".join([c for c in telefone if c.isdigit()])
            website = tags.get("website", tags.get("contact:website", ""))
            
            rua = tags.get("addr:street", "")
            numero = tags.get("addr:housenumber", "")
            bairro = tags.get("addr:suburb", "")
            endereco_completo = f"{rua}, {numero}" if rua and numero else ""
            if bairro:
                endereco_completo += f", {bairro}"
            endereco_completo += f" - {cidade}/{estado}"
            
            facebook = tags.get("contact:facebook", "")
            instagram = tags.get("contact:instagram", "")
            email = tags.get("email", tags.get("contact:email", ""))
            
            lead_data = {
                "id": i,
                "key": f"{nome}_{cidade}_{i}",
                "empresa": nome,
                "nicho": nicho,
                "categoria": tags.get("amenity", tags.get("shop", tags.get("office", categoria))),
                "estado": estado,
                "cidade": cidade,
                "endereco": endereco_completo,
                "site": website,
                "whatsapp": whatsapp,
                "telefone": telefone,
                "email": email,
                "facebook": facebook,
                "instagram": instagram,
            }
            
            analise = analisar_site(website)
            prioridade_data = calcular_prioridade_score(lead_data, analise)
            
            lead_data.update({
                "prioridade": prioridade_data["prioridade"],
                "score": prioridade_data["score"],
                "sugestoes": prioridade_data["sugestoes"],
            })
            
            leads.append(lead_data)
        
        return pd.DataFrame(leads)
        
    except Exception as e:
        st.error(f"❌ Erro: {str(e)}")
        return pd.DataFrame()


def montar_link_whatsapp(numero, mensagem):
    if not numero:
        return f"https://wa.me/?text={quote(mensagem)}"
    digits = "".join([c for c in numero if c.isdigit()])
    num_final = "55" + digits if not digits.startswith("55") else digits
    return f"https://wa.me/{num_final}?text={quote(mensagem)}"


def exportar_para_excel(leads_dict, cidade, nicho):
    """Exporta leads para Excel formatado"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Cabeçalhos
    headers = ["Empresa", "Prioridade", "Score", "Telefone", "WhatsApp", "Site", "Endereço", "Cidade", "Estado", "Sugestões"]
    ws.append(headers)
    
    # Formatar cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for lead in leads_dict.values():
        sugestoes_texto = ", ".join(lead.get("sugestoes", []))
        ws.append([
            lead.get("empresa", ""),
            lead.get("prioridade", ""),
            lead.get("score", 0),
            lead.get("telefone", ""),
            lead.get("whatsapp", ""),
            lead.get("site", ""),
            lead.get("endereco", ""),
            lead.get("cidade", ""),
            lead.get("estado", ""),
            sugestoes_texto
        ])
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 50
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ========== STATE ==========
if "df_leads" not in st.session_state:
    st.session_state.df_leads = pd.DataFrame()
if "selecionados" not in st.session_state:
    st.session_state.selecionados = {}


# ========== SIDEBAR ==========
with st.sidebar:
    st.markdown("### 🧭 Agente de Prospecção")
    st.caption("Busca inteligente com análise automática")
    st.markdown("---")
    
    st.markdown("**📍 Localização**")
    estados_ibge = buscar_estados()
    estados_opcoes = [f"{e['sigla']} - {e['nome']}" for e in estados_ibge]
    
    estado_sel = st.selectbox("Estado", estados_opcoes, index=0)
    uf = estado_sel.split(" - ")[0]
    
    cidades = buscar_cidades_por_estado(uf)
    cidade_sel = st.selectbox("Cidade", cidades, index=0)
    
    max_leads = st.slider("Quantidade de leads", 5, 50, 20, 5)
    
    st.markdown("---")
    
    st.markdown("**🎯 Nicho**")
    nichos = obter_todos_nichos()
    nicho_sel = st.selectbox("Nicho principal", nichos, index=0)
    
    categorias = obter_categorias_nicho(nicho_sel)
    categoria_sel = st.selectbox("Categoria específica", ["Todas"] + categorias, index=0)
    
    st.markdown("---")
    
    buscar_btn = st.button("🔍 Buscar Leads", type="primary", use_container_width=True)
    
    st.markdown("---")
    st.success("✅ 100% Gratuito")
    st.caption(f"🗺️ {len(estados_ibge)} estados")


# Buscar
if buscar_btn:
    st.session_state.df_leads = buscar_leads_overpass(cidade_sel, uf, max_leads, nicho_sel, categoria_sel)
    
    if not st.session_state.df_leads.empty:
        st.success(f"✅ {len(st.session_state.df_leads)} leads encontrados!")
        st.session_state.selecionados = {}

df = st.session_state.df_leads.copy()


# ========== HEADER ==========
st.title("🧭 Agente de Prospecção LP Design")

if not df.empty:
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total", len(df))
    col2.metric("Alta Prioridade", len(df[df["prioridade"].str.contains("Alta")]))
    col3.metric("Selecionados", len(st.session_state.selecionados))
    col4.metric("Score Médio", f"{df['score'].mean():.0f}")

st.markdown("---")


# ========== TABS ==========
tab_results, tab_pipeline = st.tabs(["📊 Resultados", "📌 Selecionados"])

# TAB RESULTADOS
with tab_results:
    if df.empty:
        st.info("👆 Configure os filtros e clique em **Buscar Leads**")
    else:
        # Modo de visualização
        modo = st.radio("Visualização", ["Cards", "Lista Detalhada"], horizontal=True)
        
        st.markdown("---")
        
        df = df.sort_values("score", ascending=False)
        
        if modo == "Cards":
            # MODO CARDS (3 colunas)
            for _, row in df.iterrows():
                lead_key = row["key"]
                
                with st.container(border=True):
                    col1, col2, col3 = st.columns([4, 3, 3])
                    
                    # === COLUNA 1: INFO PRINCIPAL ===
                    with col1:
                        # Logo + Nome
                        subcol_logo, subcol_nome = st.columns([1, 4])
                        with subcol_logo:
                            logo_url = buscar_logo_site(row.get("site"))
                            if logo_url:
                                st.image(logo_url, width=60)
                            else:
                                st.markdown("### 🏢")
                        
                        with subcol_nome:
                            st.markdown(f"### {row['empresa']}")
                            st.caption(f"📍 {row['endereco']}")
                        
                        # Contatos
                        if row.get("telefone"):
                            st.caption(f"📞 {row['telefone']}")
                        if row.get("whatsapp"):
                            st.caption(f"💬 WhatsApp: {row['whatsapp']}")
                        
                        # Site
                        if row.get("site"):
                            st.markdown(f"🌐 [{row['site']}]({row['site']})")
                        else:
                            st.caption("🌐 Sem site")
                        
                        # Redes sociais
                        redes = []
                        if row.get("facebook"):
                            redes.append("Facebook")
                        if row.get("instagram"):
                            redes.append("Instagram")
                        if redes:
                            st.caption(f"📱 {', '.join(redes)}")
                    
                    # === COLUNA 2: PRIORIDADE E OPORTUNIDADES ===
                    with col2:
                        st.markdown(f"## {row['prioridade']}")
                        st.metric("Score de Oportunidade", f"{row['score']}/100")
                        
                        st.markdown("**Vender:**")
                        for sug in row["sugestoes"][:4]:
                            st.caption(f"• {sug}")
                    
                    # === COLUNA 3: AÇÕES ===
                    with col3:
                        # Checkbox
                        sel = st.checkbox("📌 Selecionar", key=f"sel_{lead_key}")
                        if sel:
                            st.session_state.selecionados[lead_key] = row.to_dict()
                        else:
                            st.session_state.selecionados.pop(lead_key, None)
                        
                        # WhatsApp
                        msg = gerar_mensagem_whatsapp(row["empresa"], cidade_sel)
                        link = montar_link_whatsapp(row["whatsapp"], msg)
                        st.link_button("📲 Enviar WhatsApp", link, type="primary", use_container_width=True)
                        
                        # Relatório
                        if st.button("📄 Ver Relatório", key=f"rel_{lead_key}", use_container_width=True):
                            st.session_state[f"show_rel_{lead_key}"] = not st.session_state.get(f"show_rel_{lead_key}", False)
                
                # Relatório expandido
                if st.session_state.get(f"show_rel_{lead_key}", False):
                    st.markdown("---")
                    st.markdown("### 📋 Relatório Completo")
                    
                    col_r1, col_r2 = st.columns(2)
                    
                    with col_r1:
                        st.markdown("**📞 Contatos:**")
                        st.text(f"Telefone: {row.get('telefone', 'Não informado')}")
                        st.text(f"WhatsApp: {row.get('whatsapp', 'Não informado')}")
                        st.text(f"Email: {row.get('email', 'Não informado')}")
                        
                        st.markdown("**🌐 Presença Digital:**")
                        st.text(f"Site: {row.get('site', 'Não possui')}")
                        if row.get('facebook'):
                            st.markdown(f"[Facebook]({row['facebook']})")
                        if row.get('instagram'):
                            st.markdown(f"[Instagram]({row['instagram']})")
                        if not row.get('facebook') and not row.get('instagram'):
                            st.text("Redes Sociais: Não encontradas")
                    
                    with col_r2:
                        st.markdown("**💰 Oportunidades de Venda:**")
                        for i, sug in enumerate(row["sugestoes"], 1):
                            st.text(f"{i}. {sug}")
                        
                        st.markdown("**📊 Análise:**")
                        st.text(f"Prioridade: {row['prioridade']}")
                        st.text(f"Score: {row['score']}/100")
                        st.text(f"Categoria: {row.get('categoria', 'N/A')}")
        
        else:
            # MODO LISTA DETALHADA
            for _, row in df.iterrows():
                lead_key = row["key"]
                
                with st.container(border=True):
                    col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
                    
                    with col1:
                        st.markdown(f"**{row['empresa']}**")
                        st.caption(f"📍 {row['endereco']}")
                        if row.get("telefone"):
                            st.caption(f"📞 {row['telefone']}")
                    
                    with col2:
                        st.markdown(f"{row['prioridade']}")
                        st.caption(f"Score: {row['score']}/100")
                    
                    with col3:
                        if row.get("site"):
                            st.caption(f"🌐 {row['site']}")
                        for sug in row["sugestoes"][:2]:
                            st.caption(f"• {sug}")
                    
                    with col4:
                        sel = st.checkbox("Selecionar", key=f"sel_lista_{lead_key}")
                        if sel:
                            st.session_state.selecionados[lead_key] = row.to_dict()
                        else:
                            st.session_state.selecionados.pop(lead_key, None)
                        
                        msg = gerar_mensagem_whatsapp(row["empresa"], cidade_sel)
                        link = montar_link_whatsapp(row["whatsapp"], msg)
                        st.link_button("📲 WhatsApp", link, use_container_width=True)
                        
                        if st.button("📄", key=f"rel_lista_{lead_key}", help="Ver Relatório", use_container_width=True):
                            st.session_state[f"show_rel_{lead_key}"] = not st.session_state.get(f"show_rel_{lead_key}", False)
                
                # Relatório expandido
                if st.session_state.get(f"show_rel_{lead_key}", False):
                    st.markdown("---")
                    
                    col_r1, col_r2 = st.columns(2)
                    
                    with col_r1:
                        st.markdown("**Contatos:**")
                        st.caption(f"Tel: {row.get('telefone', 'N/A')} | Email: {row.get('email', 'N/A')}")
                        
                        st.markdown("**Digital:**")
                        st.caption(f"Site: {row.get('site', 'Não possui')}")
                        if row.get('facebook'):
                            st.caption(f"[Facebook]({row['facebook']})")
                        if row.get('instagram'):
                            st.caption(f"[Instagram]({row['instagram']})")
                    
                    with col_r2:
                        st.markdown("**Oportunidades:**")
                        for sug in row["sugestoes"]:
                            st.caption(f"• {sug}")

# TAB SELECIONADOS
with tab_pipeline:
    st.markdown("## 📌 Leads Selecionados")
    
    if not st.session_state.selecionados:
        st.info("Nenhum lead selecionado. Marque leads na aba Resultados.")
    else:
        st.success(f"✅ {len(st.session_state.selecionados)} leads selecionados")
        
        # Botão export
        try:
            excel_data = exportar_para_excel(st.session_state.selecionados, cidade_sel, nicho_sel)
            st.download_button(
                "⬇️ Baixar Excel Formatado",
                data=excel_data,
                file_name=f"leads_{cidade_sel}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        except:
            st.error("Instale openpyxl: pip install openpyxl")
        
        st.markdown("---")
        
        # Lista de selecionados
        for lead_key, lead in st.session_state.selecionados.items():
            with st.container(border=True):
                col1, col2 = st.columns([3, 1])
                
                with col1:
                    st.markdown(f"### {lead['empresa']}")
                    st.caption(f"{lead['prioridade']} | Score: {lead['score']}/100")
                    st.caption(f"📍 {lead['endereco']}")
                    st.caption(f"📞 {lead.get('telefone', 'Não informado')}")
                    
                    sugs_texto = ", ".join(lead["sugestoes"][:3])
                    st.caption(f"**Vender:** {sugs_texto}")
                
                with col2:
                    msg = gerar_mensagem_whatsapp(lead["empresa"], cidade_sel)
                    link = montar_link_whatsapp(lead["whatsapp"], msg)
                    st.link_button("📲 WhatsApp", link, type="primary", use_container_width=True)
                    
                    if st.button("🗑️ Remover", key=f"rem_{lead_key}", use_container_width=True):
                        del st.session_state.selecionados[lead_key]
                        st.rerun()
